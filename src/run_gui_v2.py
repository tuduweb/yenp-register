#/bin/python3
import os
import sys
import time
from PyQt5 import QtGui, QtWidgets, QtCore
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidget, QTableWidgetItem, QLabel, QWidget, QVBoxLayout, QTreeWidget, QTreeWidgetItem
from PyQt5.QtGui import QColor
from PyQt5.QtCore import Qt

from openpyxl import load_workbook

from core.YenpParser import YenpParser

from utils import Utils

class MainWindow(QMainWindow):

    parser:YenpParser = None

    def __init__(self):
        super().__init__()

        ########## GUI INIT ##########
        self.setWindowTitle("GUI_v2")
        self.setGeometry(100, 100, 1024, 768)

        ######### Parser Init ###########
        self.parser = YenpParser()

        # self.main_layout = self.layout() # QtWidgets.QHBoxLayout(self)
        # self.layout.setGeometry(0, 0, self.width(), self.height())
        # self.setLayout(self.layout)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # 创建一个 QHBoxLayout
        self.layout = QVBoxLayout(central_widget)

        self.start_button = QtWidgets.QPushButton(self)
        self.start_button.setText("clickButton")

        # self.start_button.clicked.connect(self.LoadProcess)

        # self.start_button.setGeometry(0, 0, 100, 20)
        self.layout.addWidget(self.start_button)
    
        self.tableWidget = QTableWidget(self)
        # self.tableWidget.setGeometry(0, 40, self.width(), self.height() - 100)
        self.layout.addWidget(self.tableWidget)


        self.labelWidget = QLabel("label", self)
        self.layout.addWidget(self.labelWidget)

        self.tableWidget.itemClicked.connect(self.handleItemClick)


        ######## QTreeWidget
        self.treeWidget = QTreeWidget(self)

        # 创建根节点
        root = QTreeWidgetItem(self.treeWidget)
        root.setText(0, "Root")
        root.setText(1, "Root Data")

        # 创建子节点
        child1 = QTreeWidgetItem(root)
        child1.setText(0, "Child 1")
        child1.setText(1, "Child 1 Data")

        child2 = QTreeWidgetItem(root)
        child2.setText(0, "Child 2")
        child2.setText(1, "Child 2 Data")

        # 在子节点上创建孙子节点
        grandchild = QTreeWidgetItem(child2)
        grandchild.setText(0, "Grandchild")
        grandchild.setText(1, "Grandchild Data")

        self.layout.addWidget(self.treeWidget)

    def handleItemClick(self, item):
        cont = item.text()
        row = item.row()+1
        column = item.column()+1
        #=======对合并的单元格取idx
        for p in self.merge_position:
            if row == p[0] and column == p[1]:
                row = row + (p[2]-p[0])
                break
        #=======对合并的单元格取idx
        # self.comboBox_x.addItem(str(row))
        # self.comboBox_y.addItem(str(column))
        # self.comboBox_r1.addItem(str(row+1))
        print(str(row), str(column), str(row+1))

    def InitTable(self):
        self.tableWidget.setColumnCount(10)  # 设置表格的列数

        # self.UpdateTableHeader()
        self.AddTableRowItem()

        self.ValidateTable()

    def UpdateTableHeader(self, _headerTitleLabels):
        self.tableWidget.setHorizontalHeaderLabels(_headerTitleLabels)

        pass

    def AddTableRowItem(self):
        # self.tableWidget.
        # self.tableWidget.setItem(0, 0, QTableWidgetItem("Cell 1"))
        # self.tableWidget.setItem(0, 1, QTableWidgetItem("Cell 2"))
        # self.tableWidget.setItem(0, 2, QTableWidgetItem("Cell 3"))

        file_path = "./examples/reg_table1.xlsx"
        wb = load_workbook(filename=file_path)
        sheet_names = wb.sheetnames        
        ws = wb[sheet_names[0]]

        title_row_number = 2  # 指定要打印的行号

        max_row, max_colum = ws.max_row, ws.max_column
        self.tableWidget.setRowCount(max_row - title_row_number)  # 设置表格的列数

        ########## merge cells ##########
        merge_idx = ws.merged_cells
        merge_idx = Utils.get_merge_cell_list(merge_idx)

        # for handle click event
        self.merge_position = []

        for i in range(len(merge_idx)):
            m_idx = merge_idx[i]
            self.tableWidget.setSpan(m_idx[0] - 1 - title_row_number,  m_idx[1]-1, m_idx[2]-m_idx[0]+1, m_idx[3]-m_idx[1]+1)
            self.merge_position.append([m_idx[0],m_idx[1],m_idx[2]])#[x1,y1,range]

        ########## Sheet Titles ##########
        sheet_row_data = [item for item in ws.iter_rows(min_row=title_row_number, max_row=title_row_number, values_only=True)]
        sheet_title_list = [item for item in sheet_row_data[0]]
        # print(sheet_title_list)
        # print("sheet_title row=%d" % row_number, sheet_row_data, sheet_title_list)
        self.UpdateTableHeader(sheet_title_list)

        self.yenpRegData = {}

        blockItems = [(idx,item.value) for idx, item in enumerate(ws['A'][title_row_number:]) if item.value is not None]
        print(blockItems)

        for rowWrapper in ws.iter_rows(min_row=title_row_number):
            # YenpRegItem

            # decode item_row            
            for cell in rowWrapper:
                # print(cell, type(cell))
                # print("Cell: ({}, {}) - Value: {}".format(cell.row, cell.column, cell.value))

                item = QTableWidgetItem()
                item.setData(Qt.UserRole, "UserData")
                if cell.value is not None:
                    item.setText(str(cell.value))
                # else:
                #     item = QTableWidgetItem()

                self.tableWidget.setItem(cell.row - 1 - title_row_number, cell.column - 1, item)


    def ValidateTable(self):
        #for column in range(self.tableWidget.columnCount()):
            # item = QTableWidgetItem("Cell {}".format(column + 1))
        #    item.setBackground(QColor(100, 0, 0))  # 设置背景颜色为红色
        #    self.tableWidget.setItem(0, column, item)
        pass

if __name__ == "__main__":
    app = QApplication(sys.argv)
    main_window = MainWindow()
    main_window.show()
    
    main_window.InitTable()

    sys.exit(app.exec_())