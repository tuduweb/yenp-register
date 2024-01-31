#/bin/python3
import os
import sys
import time
from PyQt5 import QtGui, QtWidgets, QtCore
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidget, QTableWidgetItem
from openpyxl import load_workbook

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.activate_file = [None,None]

        self.setWindowTitle("TableWidget Example")
        self.setGeometry(100, 100, 1024, 768)

        self.main_layout = self.layout() # QtWidgets.QHBoxLayout(self)
        # self.layout.setGeometry(0, 0, self.width(), self.height())
        # self.setLayout(self.layout)

        self.start_button = QtWidgets.QPushButton(self)
        self.start_button.setText("clickButton")

        self.start_button.clicked.connect(self.LoadProcess)

        # self.start_button.setGeometry(0, 0, 100, 20)
        self.main_layout.addWidget(self.start_button)
    
        self.tableWidget = QTableWidget(self)
        self.tableWidget.setGeometry(0, 40, self.width(), self.height() - 40)
        self.main_layout.addWidget(self.tableWidget)

        # Set the column count and headers
        # self.tableWidget.setColumnCount(3)
        # self.tableWidget.setHorizontalHeaderLabels(["Name", "Age", "Gender"])



        # Add some sample data
        # self.add_data("John Doe", 25, "Male")
        # self.add_data("Jane Smith", 30, "Female")
        # self.add_data("Bob Johnson", 35, "Male")

    def add_data(self, name, age, gender):
        row_count = self.tableWidget.rowCount()
        self.tableWidget.insertRow(row_count)

        self.tableWidget.setItem(row_count, 0, QTableWidgetItem(name))
        self.tableWidget.setItem(row_count, 1, QTableWidgetItem(str(age)))
        self.tableWidget.setItem(row_count, 2, QTableWidgetItem(gender))

    def LoadProcess(self):
        # self.clearcontext_all()
        # if self.comboBoxfiletype.currentIndex() == 1:  # xls
        #     QMessageBox.about(self, "hi,兰神", '不支持 xls 格式文件')
        # elif self.comboBoxfiletype.currentIndex() == 0:  # xlsx
        #     items = self.listWidget.selectedItems()
        #     if len(items) == 0:
        #         QMessageBox.about(self, "hi,兰神", '请先选择文件')
        #     else:
        items = ["./examples/reg_table1.xlsx"]
        self.infos = {}
        for i in list(items):
            file_path = str(i)
            wb = load_workbook(filename=file_path)
            name = os.path.split(file_path)[-1]

            sheet_names = wb.sheetnames

            sheets_dict = {}
            for s in sheet_names:
                sheets_dict[s] = []
            self.infos[name] = {'path':file_path,'sheet_names':sheets_dict}
            wb.close()
        #for k in self.infos.keys():
        #    self.comboBox_wb.addItem(k)
        #k = self.comboBox_wb.itemText(0)
        #sheets = list(self.infos[k]['sheet_names'].keys())
        #for s in sheets:
        #    self.comboBox_ws.addItem(s)

        for k in self.infos.keys():
            self.activate_file[0] = self.infos[k]['path']
            self.activate_file[1] = list(self.infos[k]['sheet_names'].keys())[0]

        self.show_excel()

    def show_excel(self):
        self.merge_position = []
        path = self.activate_file[0]
        sheetname = self.activate_file[1]
        wb = load_workbook(filename=path)
        ws = wb[sheetname]
        num_row = ws.max_row
        num_column = ws.max_column
        self.tableWidget.setColumnCount(num_column)
        self.tableWidget.setRowCount(num_row)



        #======合并单元格=======
        merge_idx = ws.merged_cells
        merge_idx = get_merge_cell_list(merge_idx)

        for i in range(len(merge_idx)):
            m_idx = merge_idx[i]
            self.tableWidget.setSpan(m_idx[0]-1, m_idx[1]-1, m_idx[2]-m_idx[0]+1, m_idx[3]-m_idx[1]+1)
            self.merge_position.append([m_idx[0],m_idx[1],m_idx[2]])#[x1,y1,range]
        #======合并单元格=======

        #======单元格大小=======
        for i in range(1,num_row+1):
            h = ws.row_dimensions[i].height
            if h is not None:
                self.tableWidget.setRowHeight(i-1,h)

        for i in range(1,num_row+1):
            #self.comboBox_r2.addItem(str(num_row-i+1))
            row_sizes = []
            for j in range(1,num_column+1):
                cell = ws.cell(row=i, column=j)
                if cell.value is not None:
                    item = QTableWidgetItem(str(cell.value))
                    assign_style_qt(item,cell)
                else:
                    item = QTableWidgetItem()
                self.tableWidget.setItem(i-1, j-1, item)

        _headerTitleLabels = []
        for idx in range(1, num_column + 1):
            cell = ws.cell(row=2, column=idx)
            _headerTitleLabels.append(str(cell.value))
        self.tableWidget.setHorizontalHeaderLabels(_headerTitleLabels)



from openpyxl import Workbook, load_workbook,styles
import os
import copy
import math
from openpyxl.utils import get_column_letter, column_index_from_string
from PyQt5.QtGui import QFont
from PyQt5.QtCore import Qt

def get_merge_cell_list(merge_idx):
    merge_idx = list(merge_idx)
    merge_list = []
    for i in range(len(merge_idx)):
        merge = merge_idx[i]
        row_min, row_max, col_min, col_max = merge.min_row, merge.max_row, merge.min_col, merge.max_col
        # merge_list.append([row_min, row_max, col_min, col_max])
        merge_list.append([row_min, col_min, row_max, col_max])
    return merge_list

def assign_style_qt(target_cell,source_cell):
    #字体，大小，颜色，加粗
    font = QFont()   #实例化字体对象
    font.setFamily(source_cell.font.name)  #字体
    font.setBold(source_cell.font.bold)  #加粗
    font.setPointSize(source_cell.font.size)   #字体大小
    target_cell.setFont(font)
    #居中
    target_cell.setTextAlignment(Qt.AlignCenter | Qt.AlignCenter)
    #背景
    # print(source_cell.fill.bgColor.rgb)
    # target_cell.setBackground(QtGui.QColor(100,100,150))

if __name__ == "__main__":
    app = QApplication(sys.argv)
    main_window = MainWindow()
    main_window.show()
    sys.exit(app.exec_())
