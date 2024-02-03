import os
from openpyxl import load_workbook
# from ExcelParser import ExcelParser

def get_merge_cell_list(merge_idx):
    merge_idx = list(merge_idx)
    merge_list = []
    for i in range(len(merge_idx)):
        merge = merge_idx[i]
        row_min, row_max, col_min, col_max = merge.min_row, merge.max_row, merge.min_col, merge.max_col
        # merge_list.append([row_min, row_max, col_min, col_max])
        merge_list.append([row_min, col_min, row_max, col_max])
    return merge_list


class YenpParser(object):
    infos = None

    def __init__(self) -> None:
        self.infos = {}

    def LoadExcel(self, file_path) -> dict:
        wb = load_workbook(filename=file_path)
        name = os.path.split(file_path)[-1]
        sheet_names = wb.sheetnames
        sheets_dict = {}

        for s in sheet_names:
            sheets_dict[s] = []
        self.infos[name] = {'path':file_path,'sheet_names':sheets_dict}
        wb.close()

        print(self.infos)

        self.activate_file = {}
        for k in self.infos.keys():
            self.activate_file = self.infos[k]

        return self.infos
    
    def LoadExcelSheet(self, idx = 0):
        print(list(self.activate_file["sheet_names"].keys())[0])
        _path, _sheetname = self.activate_file["path"], list(self.activate_file["sheet_names"].keys())[0]
        wb = load_workbook(filename=_path)
        ws = wb[_sheetname]
        
        max_row, max_colum = ws.max_row, ws.max_column
        
        row_number = 1  # 指定要打印的行号
        row_data = ws[row_number]

        # 打印指定行的数据
        for cell in row_data:
            print(cell.value)

        ### Excel Parser

        sheet_data = {
            "base_addr_index": 1,
            "title_index": 2,
            "register_start_index": 3
        }

        ## Find Base Addr Row
        row_number = 1  # 指定要打印的行号
        row_data = ws.iter_rows(min_row=row_number, max_row=row_number, values_only=True)
        print("base_addr row=%d" % row_number, list(row_data))

        # Find Title Row
        row_number = 2  # 指定要打印的行号
        row_data = ws.iter_rows(min_row=row_number, max_row=row_number, values_only=True)
        print("sheet_title row=%d" % row_number, list(row_data))

        # Other Row -> data rows
        for row in ws.iter_rows(min_row=row_number + 1, values_only=True):
            print(row)
        # for j in range(1, num_column + 1):

        # merge cells
        merge_idx = ws.merged_cells
        print("merge_idx", type(merge_idx), merge_idx)

        merge_idx_list = get_merge_cell_list(merge_idx)
        print("merge_idx_list", merge_idx_list)

        print("sheet_data", sheet_data)
    
if __name__ == "__main__":
    print("hello world")
    import argparse
    _arg_parser = argparse.ArgumentParser(description='YenpParser function!~')
    _arg_parser.add_argument('-i', '--input_file', help='文件路径', default="./examples/reg_table1.xlsx")
    _arg_parser.add_argument('-o', '--output_file', help='输出文件的路径', required=False)
    args = _arg_parser.parse_args()


    # Check File Path Valid

    parser = YenpParser()
    parser.LoadExcel(args.input_file)
    parser.LoadExcelSheet(0)